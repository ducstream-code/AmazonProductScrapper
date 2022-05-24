import openpyxl
import sys
from PySide6 import *
from PySide6 import QtWidgets
from PySide6.QtCore import *
from PySide6.QtWidgets import *
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)

        self.setMinimumSize(QSize(350, 300))
        self.setWindowTitle("PyQt Line Edit example (textfield) - pythonprogramminglanguage.com")

        self.objectColumn = QLabel(self)
        self.objectColumn.setText("Colonne des objets: ")
        self.objectColumn.resize(120, 10)
        self.objectColumn.move(20, 30)
        self.objectColumnInput = QLineEdit(self)
        self.objectColumnInput.move(142, 20)
        self.objectColumnInput.resize(32, 32)

        self.writingColumn = QLabel(self)
        self.writingColumn.setText("Colonne d\'écriture:")
        self.writingColumn.resize(120, 10)
        self.writingColumn.move(20, 70)
        self.writingColumnInput = QLineEdit(self)
        self.writingColumnInput.move(140, 60)
        self.writingColumnInput.resize(32, 32)

        self.startLine = QLabel(self)
        self.startLine.setText("Ligne de départ:")
        self.startLine.resize(120, 15)
        self.startLine.move(20, 110)
        self.startLineInput = QLineEdit(self)
        self.startLineInput.move(120, 100)
        self.startLineInput.resize(32, 32)

        self.fileName = QLabel(self)
        self.fileName.setText("Nom du fichier excel:")
        self.fileName.resize(130, 15)
        self.fileName.move(20, 150)
        self.fileNameInput = QLineEdit(self)
        self.fileNameInput.move(160, 140)
        self.fileNameInput.resize(100, 32)

        self.lineNumber = QLabel(self)
        self.lineNumber.setText("Nombre de lignes:")
        self.lineNumber.resize(130, 15)
        self.lineNumber.move(20, 190)
        self.lineNumberInput = QLineEdit(self)
        self.lineNumberInput.move(160, 180)
        self.lineNumberInput.resize(100, 32)

        pybutton = QPushButton('OK', self)
        pybutton.clicked.connect(self.clickMethod)
        pybutton.resize(200, 32)
        pybutton.move(20, 220)

        self.prog_bar = QProgressBar(self)
        self.prog_bar.setGeometry(50, 50, 300, 30)
        self.prog_bar.move(20, 265)

    def updateStatus(self):
        value = self.prog_bar.value()
        self.prog_bar.setValue(value + 1)

    def clickMethod(self):
        # defining the range of the loading bar depending on the number of line and the starting line
        self.prog_bar.setRange(0, int(self.lineNumberInput.text()) - int(self.startLineInput.text()))
        # Variable section:

        # The xlsx sheet you want to open
        start = int(self.startLineInput.text())

        fileName = self.fileNameInput.text()

        # The column where object names are written
        ObjectColumn = int(self.objectColumnInput.text())

        # Number of row in your file
        rowNumber = int(self.lineNumberInput.text())

        # The column where you want to write the fetched data
        writingColumn = int(self.writingColumnInput.text())

        # Loop between 1 and the number of row in the sheet
        for i in range(start, rowNumber):
            QCoreApplication.processEvents()
            wb = openpyxl.load_workbook(fileName)
            sh = wb.active
            cell = sh.cell(row=i, column=ObjectColumn)
            print(cell.value)
            objects = cell.value
            # try is here in case the page doesn't have any result for the need classname
            try:
                op = webdriver.ChromeOptions();
                #uncomment the line below to not display chrome windows (can bug sometimes)
                #op.add_argument('headless')
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=op)
                # Define the link to search (passing the object in the link)
                driver.get(
                    "https://www.amazon.fr/s?k=" + objects + "&rh=n%3A838343031&__mk_fr_FR=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss")

                # The classname you want to get
                item = driver.find_element(By.CLASS_NAME, "a-price-whole")
                print(item.text)
                sh.cell(row=i, column=writingColumn).value = item.text
                wb.save(fileName)
                self.updateStatus()

            except Exception as e:
                print(e)
                wb.save(fileName)
                self.updateStatus()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    mainWin = MainWindow()
    mainWin.show()
    sys.exit(app.exec())
